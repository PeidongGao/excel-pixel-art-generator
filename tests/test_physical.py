import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from excel_pixel_art.physical import (
    image_to_physical_excel,
    image_to_physical_masks,
    image_to_physical_pdf,
    main,
    parse_resolution,
)


class PhysicalLayerTest(unittest.TestCase):
    def test_physical_workbook_has_independent_palette_tiles_and_aligned_masks(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "physical.xlsx"

            image = Image.new("RGBA", (4, 2))
            colors = [
                (255, 0, 0, 255),
                (0, 255, 0, 255),
                (0, 0, 255, 255),
                (255, 255, 0, 255),
            ]
            for row in range(2):
                for column in range(4):
                    image.putpixel((column, row), colors[column])
            image.save(image_path)

            image_to_physical_excel(
                image_path,
                output_path,
                canvas_size="letter",
                resolution=(3, 2),
                orientation="portrait",
                material_color_count=2,
                cell_size=4.0,
                poster_pages=(2, 1),
                generate_color_masks=True,
                max_color_masks=2,
            )

            workbook = load_workbook(output_path)
            template = workbook["Print Template"]
            palette = workbook["Material Palette"]
            first_mask = workbook["Material Mask 001"]

            self.assertEqual(
                workbook.sheetnames,
                [
                    "Print Reference",
                    "Print Template",
                    "Material Palette",
                    "Material Mask 001",
                    "Material Mask 002",
                ],
            )
            self.assertEqual((template.max_column, template.max_row), (3, 2))
            self.assertEqual(template.page_setup.paperSize, 1)
            self.assertEqual(template.page_setup.orientation, "portrait")
            self.assertEqual(template.page_setup.fitToWidth, 2)
            self.assertEqual(len(template.col_breaks.brk), 1)
            self.assertEqual(palette.max_row - 1, 2)
            self.assertEqual(first_mask.oddHeader.center.text, f"Color 1 - {palette['B2'].value}")
            self.assertGreater(
                sum(cell.fill.fill_type == "solid" for row in first_mask.iter_rows() for cell in row),
                0,
            )

    def test_physical_defaults_are_a4_128_square_and_48_colors(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "physical.xlsx"
            Image.new("RGBA", (4, 2), (20, 40, 60, 255)).save(image_path)

            image_to_physical_excel(image_path, output_path)

            workbook = load_workbook(output_path)
            template = workbook["Print Template"]
            self.assertEqual((template.max_column, template.max_row), (128, 128))
            self.assertEqual(template.page_setup.paperSize, 9)
            self.assertLessEqual(workbook["Material Palette"].max_row - 1, 48)

    def test_fixed_material_palettes_include_names_and_references(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "lego.xlsx"
            Image.new("RGBA", (4, 2), (200, 20, 20, 255)).save(image_path)

            image_to_physical_excel(image_path, output_path, resolution=(4, 2), palette_mode="lego")

            palette = load_workbook(output_path)["Material Palette"]
            self.assertEqual(palette["E1"].value, "Material")
            self.assertTrue(palette["E2"].value)
            self.assertEqual(palette["G2"].value, "https://www.bricklink.com/catalogColors.asp")

            liquitex_path = directory_path / "liquitex.xlsx"
            image_to_physical_excel(image_path, liquitex_path, resolution=(4, 2), palette_mode="liquitex_basics_24")
            liquitex_palette = load_workbook(liquitex_path)["Material Palette"]
            used_materials = {
                liquitex_palette.cell(row=row, column=5).value
                for row in range(2, liquitex_palette.max_row + 1)
            }
            self.assertTrue(used_materials <= {
                "Bright Aqua Green", "Burnt Sienna", "Burnt Umber", "Cadmium Orange Hue",
                "Cadmium Red Deep Hue", "Cadmium Red Medium Hue", "Cadmium Yellow Deep Hue",
                "Cadmium Yellow Medium Hue", "Dioxazine Purple", "Iridescent Gold",
                "Iridescent Silver", "Ivory Black", "Mars Black", "Medium Magenta",
                "Naphthol Crimson", "Permanent Green Light", "Permanent Hooker's Green",
                "Permanent Light Blue", "Phthalo Blue", "Primary Blue", "Primary Yellow",
                "Titanium White", "Ultramarine Blue", "Unbleached Titanium",
            })

    def test_printable_pdf_and_mask_zip_outputs(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            pdf_path = directory_path / "printable.pdf"
            masks_path = directory_path / "masks.zip"
            Image.new("RGBA", (4, 2), (200, 20, 20, 255)).save(image_path)

            image_to_physical_pdf(image_path, pdf_path, resolution=(4, 2), poster_pages=(2, 1))
            image_to_physical_masks(image_path, masks_path, resolution=(4, 2), palette_mode="liquitex_basics_24")

            self.assertTrue(pdf_path.read_bytes().startswith(b"%PDF"))
            self.assertTrue(masks_path.read_bytes().startswith(b"PK"))

    def test_rejects_invalid_physical_options(self):
        with self.assertRaises(ValueError):
            image_to_physical_excel("image.png", "output.xlsx", poster_pages=(0, 1))
        with self.assertRaises(ValueError):
            image_to_physical_excel("image.png", "output.xlsx", max_color_masks=0)
        with self.assertRaises(ValueError):
            image_to_physical_excel("image.png", "output.xlsx", resolution=(0, 10))
        with self.assertRaises(ValueError):
            image_to_physical_excel("image.png", "output.xlsx", material_color_count=1)
        with self.assertRaises(ValueError):
            image_to_physical_excel("image.png", "output.xlsx", cell_size=0)

    def test_physical_cli(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "physical.xlsx"
            Image.new("RGBA", (2, 2), (20, 40, 60, 255)).save(image_path)

            result = main([str(image_path), "-o", str(output_path), "--resolution", "4x3"])

            self.assertEqual(result, 0)
            self.assertTrue(output_path.exists())
            self.assertEqual(parse_resolution("4x3"), (4, 3))


if __name__ == "__main__":
    unittest.main()
