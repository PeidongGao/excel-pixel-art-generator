import io
import unittest
import zipfile

from openpyxl import load_workbook
from PIL import Image
from streamlit.testing.v1 import AppTest

from excel_pixel_art.streamlit_app import (
    _build_digital_workbook,
    _build_physical_outputs,
    _build_physical_workbook,
)


class UploadedImage:
    name = "source.png"

    def __init__(self):
        image = Image.new("RGBA", (4, 2), (20, 40, 60, 255))
        content = io.BytesIO()
        image.save(content, format="PNG")
        self._content = content.getvalue()

    def getvalue(self):
        return self._content


class StreamlitAppTest(unittest.TestCase):
    def test_layer_defaults_and_separate_generation_controls(self):
        app = AppTest.from_file("excel_pixel_art/streamlit_app.py").run(timeout=20)

        self.assertEqual(len(app.exception), 0)
        self.assertEqual([header.value for header in app.header], ["Digital Layer", "Physical Layer"])
        self.assertEqual([button.label for button in app.button], [
            "Generate Digital workbook",
            "Generate Physical outputs",
        ])

        number_inputs = {number_input.label: number_input.value for number_input in app.number_input}
        sliders = {slider.label: slider.value for slider in app.slider}
        selectboxes = {selectbox.label: selectbox.value for selectbox in app.selectbox}

        self.assertEqual(selectboxes["Paper size"], "a4")
        self.assertEqual(number_inputs["Width cells"], 32)
        self.assertEqual(number_inputs["Height cells"], 32)
        self.assertEqual(sliders["Indexed colors"], 24)
        self.assertEqual(selectboxes["Physical paper size"], "a4")
        self.assertEqual(number_inputs["Physical width cells"], 32)
        self.assertEqual(number_inputs["Physical height cells"], 32)
        self.assertEqual(sliders["Material Palette colors"], 24)
        self.assertIn("Status: **Valid**", app.success[0].value)
        self.assertIn("Use, Privacy, and Liability", [item.value for item in app.subheader])
        self.assertIn("Important information", [item.label for item in app.expander])

    def test_invalid_poster_split_displays_divisibility_reason(self):
        app = AppTest.from_file("excel_pixel_art/streamlit_app.py").run(timeout=20)

        app.checkbox[2].set_value(True)
        app.number_input[6].set_value(3)
        app.run(timeout=20)

        self.assertEqual(len(app.error), 1)
        self.assertIn("Status: **Invalid**", app.error[0].value)
        self.assertIn("Width 32 is not divisible by 3.", app.error[0].value)

    def test_poster_split_requires_exact_resolution_for_verification(self):
        app = AppTest.from_file("excel_pixel_art/streamlit_app.py").run(timeout=20)

        app.checkbox[1].set_value(False)
        app.checkbox[2].set_value(True)
        app.run(timeout=20)

        warning_text = "\n".join(item.value for item in app.warning)
        self.assertIn("Status: Pending", warning_text)
        self.assertIn("Enable exact physical resolution", warning_text)

    def test_digital_and_physical_builders_create_separate_outputs(self):
        uploaded_file = UploadedImage()

        digital_bytes, digital_name = _build_digital_workbook(
            uploaded_file=uploaded_file,
            max_size=128,
            cell_size=3.0,
            color_count=48,
            canvas_size="a4",
            resolution=(8, 6),
            orientation="landscape",
            fit="contain",
            background_color="FFFFFF",
        )
        physical_bytes, physical_name = _build_physical_workbook(
            uploaded_file=uploaded_file,
            max_size=128,
            cell_size=3.0,
            material_color_count=24,
            canvas_size="letter",
            resolution=(5, 7),
            orientation="portrait",
            fit="cover",
            background_color="EEEEEE",
            poster_pages=(2, 2),
            generate_color_masks=True,
            max_color_masks=1,
        )

        digital = load_workbook(io.BytesIO(digital_bytes))
        physical = load_workbook(io.BytesIO(physical_bytes))

        self.assertEqual(digital_name, "source_digital.xlsx")
        self.assertEqual(physical_name, "source_physical.xlsx")
        self.assertEqual(digital.sheetnames, ["Reference", "Template", "Color Index"])
        self.assertEqual(
            physical.sheetnames,
            ["Print Reference", "Print Template", "Material Palette", "Material Mask 001"],
        )
        self.assertEqual((digital["Template"].max_column, digital["Template"].max_row), (8, 6))
        self.assertEqual((physical["Print Template"].max_column, physical["Print Template"].max_row), (5, 7))

    def test_physical_output_builder_returns_selected_independent_files(self):
        archive_bytes, archive_name = _build_physical_outputs(
            uploaded_file=UploadedImage(),
            output_workbook=True,
            output_pdf=True,
            output_masks=True,
            max_size=128,
            cell_size=3.0,
            material_color_count=24,
            palette_mode="liquitex_basics_24",
            canvas_size="a4",
            resolution=(8, 8),
            orientation="portrait",
            fit="contain",
            background_color="FFFFFF",
            poster_pages=(2, 2),
            generate_color_masks=True,
            max_color_masks=2,
        )

        self.assertEqual(archive_name, "source_physical_outputs.zip")
        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
            self.assertEqual(
                set(archive.namelist()),
                {"source_physical.xlsx", "source_printable.pdf", "source_masks.zip"},
            )


if __name__ == "__main__":
    unittest.main()
