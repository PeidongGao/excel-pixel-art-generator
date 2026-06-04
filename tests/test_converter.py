import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from excel_pixel_art.converter import image_to_excel


class ConverterTest(unittest.TestCase):
    def test_image_to_excel_writes_pixel_colors(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "pixel_art.xlsx"

            image = Image.new("RGBA", (2, 1))
            image.putpixel((0, 0), (255, 0, 0, 255))
            image.putpixel((1, 0), (0, 128, 255, 255))
            image.save(image_path)

            result = image_to_excel(image_path, output_path, max_size=10)

            self.assertEqual(result, output_path)
            workbook = load_workbook(output_path)
            sheet = workbook.active
            self.assertEqual(workbook.sheetnames, ["Reference", "Template", "Color Index"])
            self.assertEqual(sheet["A1"].fill.fgColor.rgb, "00FF0000")
            self.assertEqual(sheet["B1"].fill.fgColor.rgb, "000080FF")

    def test_workbook_includes_template_and_color_index(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "paint_by_number.xlsx"

            image = Image.new("RGBA", (3, 1))
            image.putpixel((0, 0), (255, 0, 0, 255))
            image.putpixel((1, 0), (0, 0, 255, 255))
            image.putpixel((2, 0), (255, 0, 0, 255))
            image.save(image_path)

            image_to_excel(image_path, output_path, max_size=10)

            workbook = load_workbook(output_path)
            template = workbook["Template"]
            color_index = workbook["Color Index"]

            self.assertEqual(color_index["A1"].value, "Number")
            self.assertEqual(color_index["B1"].value, "Hex Code")
            self.assertEqual(color_index["D1"].value, "Cells")
            self.assertEqual(color_index["A2"].value, 1)
            self.assertEqual(color_index["B2"].value, "#FF0000")
            self.assertEqual(color_index["D2"].value, 2)
            self.assertEqual(color_index["A3"].value, 2)
            self.assertEqual(color_index["B3"].value, "#0000FF")
            self.assertEqual(color_index["D3"].value, 1)
            self.assertEqual(template["A1"].value, 1)
            self.assertEqual(template["B1"].value, 2)
            self.assertEqual(template["C1"].value, 1)
            self.assertEqual(template["A1"].fill.fill_type, None)
            self.assertEqual(template["A1"].font.color.rgb, "00A6A6A6")
            self.assertEqual(template["A1"].border.left.style, None)
            self.assertEqual(template.print_options.gridLines, False)
            self.assertEqual(template.sheet_view.showGridLines, True)

    def test_color_count_limits_indexed_palette(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "limited_palette.xlsx"

            image = Image.new("RGBA", (4, 1))
            for column, color in enumerate(
                [
                    (255, 0, 0, 255),
                    (0, 255, 0, 255),
                    (0, 0, 255, 255),
                    (255, 255, 0, 255),
                ]
            ):
                image.putpixel((column, 0), color)
            image.save(image_path)

            image_to_excel(image_path, output_path, color_count=2)

            workbook = load_workbook(output_path)
            color_index = workbook["Color Index"]
            self.assertLessEqual(color_index.max_row - 1, 2)

    def test_transparent_pixels_are_left_unfilled(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "transparent.png"
            output_path = directory_path / "pixel_art.xlsx"

            image = Image.new("RGBA", (1, 1), (255, 0, 0, 0))
            image.save(image_path)

            image_to_excel(image_path, output_path)

            workbook = load_workbook(output_path)
            sheet = workbook.active
            self.assertEqual(sheet["A1"].fill.fill_type, None)

    def test_canvas_preset_creates_fixed_a4_landscape_grid(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "a4.xlsx"

            Image.new("RGBA", (200, 100), (20, 40, 60, 255)).save(image_path)

            image_to_excel(
                image_path,
                output_path,
                max_size=20,
                canvas_size="a4",
                orientation="landscape",
            )

            workbook = load_workbook(output_path)
            sheet = workbook.active
            self.assertEqual(sheet.max_column, 20)
            self.assertEqual(sheet.max_row, 14)
            self.assertEqual(sheet.page_setup.orientation, "landscape")
            self.assertEqual(sheet.page_setup.paperSize, 9)

    def test_resolution_creates_exact_grid(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "resolution.xlsx"

            Image.new("RGBA", (20, 10), (20, 40, 60, 255)).save(image_path)

            image_to_excel(image_path, output_path, resolution=(12, 7))

            workbook = load_workbook(output_path)
            sheet = workbook.active
            self.assertEqual(sheet.max_column, 12)
            self.assertEqual(sheet.max_row, 7)

    def test_physical_output_uses_independent_resolution_palette_and_aligned_masks(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "print_mode.xlsx"

            image = Image.new("RGBA", (4, 2))
            colors = [
                (255, 0, 0, 255),
                (0, 255, 0, 255),
                (0, 0, 255, 255),
                (255, 255, 0, 255),
            ]
            for row in range(2):
                for column in range(4):
                    image.putpixel((column, row), colors[column])
            image.save(image_path)

            image_to_excel(
                image_path,
                output_path,
                resolution=(4, 2),
                color_count=4,
                include_excel_output=True,
                physical_output=True,
                poster_pages=(2, 1),
                generate_color_masks=True,
                max_color_masks=2,
                physical_resolution=(3, 2),
                material_color_count=2,
                physical_cell_size=4.0,
            )

            workbook = load_workbook(output_path)
            reference = workbook["Reference"]
            template = workbook["Template"]
            color_index = workbook["Color Index"]
            print_template = workbook["Print Template"]
            material_palette = workbook["Material Palette"]
            first_mask = workbook["Material Mask 001"]

            self.assertEqual(
                workbook.sheetnames,
                [
                    "Reference",
                    "Template",
                    "Color Index",
                    "Print Reference",
                    "Print Template",
                    "Material Palette",
                    "Material Mask 001",
                    "Material Mask 002",
                ],
            )
            self.assertEqual(reference.page_setup.fitToWidth, 1)
            self.assertEqual(template.page_setup.fitToHeight, 1)
            self.assertEqual(template.max_column, 4)
            self.assertEqual(template.max_row, 2)
            self.assertEqual(print_template.max_column, 3)
            self.assertEqual(print_template.max_row, 2)
            self.assertEqual(print_template.page_setup.fitToWidth, 2)
            self.assertEqual(len(print_template.col_breaks.brk), 1)
            self.assertEqual(len(print_template.row_breaks.brk), 0)
            self.assertEqual(color_index.max_row - 1, 4)
            self.assertEqual(material_palette.max_row - 1, 2)
            self.assertNotEqual(list(color_index.values), list(material_palette.values))
            self.assertEqual(first_mask.oddHeader.center.text, f"Color 1 - {material_palette['B2'].value}")
            self.assertEqual(first_mask.sheet_view.showGridLines, False)
            self.assertEqual(first_mask.print_options.gridLines, False)
            filled_cells = sum(
                cell.fill.fill_type == "solid"
                for row in first_mask.iter_rows()
                for cell in row
            )
            self.assertGreater(filled_cells, 0)

    def test_physical_output_can_be_generated_without_excel_output(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "physical_only.xlsx"

            Image.new("RGBA", (2, 2), (255, 0, 0, 255)).save(image_path)

            image_to_excel(
                image_path,
                output_path,
                resolution=(2, 2),
                include_excel_output=False,
                physical_output=True,
                physical_resolution=(3, 4),
                material_color_count=2,
            )

            workbook = load_workbook(output_path)
            self.assertEqual(
                workbook.sheetnames,
                ["Print Reference", "Print Template", "Material Palette"],
            )
            self.assertEqual(workbook["Print Template"].max_column, 3)
            self.assertEqual(workbook["Print Template"].max_row, 4)

    def test_physical_canvas_settings_are_independent_from_digital_canvas(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            image_path = directory_path / "source.png"
            output_path = directory_path / "independent_canvases.xlsx"

            Image.new("RGBA", (20, 10), (20, 40, 60, 255)).save(image_path)

            image_to_excel(
                image_path,
                output_path,
                canvas_size="a4",
                resolution=(12, 7),
                orientation="landscape",
                include_excel_output=True,
                physical_output=True,
                physical_canvas_size="letter",
                physical_resolution=(5, 9),
                physical_orientation="portrait",
            )

            workbook = load_workbook(output_path)
            digital = workbook["Template"]
            physical = workbook["Print Template"]

            self.assertEqual((digital.max_column, digital.max_row), (12, 7))
            self.assertEqual((physical.max_column, physical.max_row), (5, 9))
            self.assertEqual(digital.page_setup.paperSize, 9)
            self.assertEqual(digital.page_setup.orientation, "landscape")
            self.assertEqual(physical.page_setup.paperSize, 1)
            self.assertEqual(physical.page_setup.orientation, "portrait")

    def test_rejects_invalid_size_options(self):
        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", max_size=0)

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", cell_size=0)

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", canvas_size="poster")

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", fit="stretch")

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", resolution=(0, 10))

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", color_count=1)

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", poster_pages=(0, 1))

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", max_color_masks=0)

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", physical_resolution=(0, 10))

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", material_color_count=1)

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", physical_cell_size=0)

        with self.assertRaises(ValueError):
            image_to_excel("image.png", "output.xlsx", include_excel_output=False)


if __name__ == "__main__":
    unittest.main()
